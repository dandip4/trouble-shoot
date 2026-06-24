from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExportService:
    HEADER_FILL = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    def _apply_header_style(self, ws, row=1):
        for cell in ws[row]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_width(self, ws):
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            column_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[column_letter].width = min(max(length + 2, 10), 50)

    def _apply_alternate_rows(self, ws):
        for row in ws.iter_rows(min_row=2):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = PatternFill(start_color="F3F6FB", end_color="F3F6FB", fill_type="solid")

    def generate_report(self, rows, summary, clustering_info, periode_label):
        wb = Workbook()
        data_sheet = wb.active
        data_sheet.title = "Data Troubleshoot"

        headers = [
            "No SPK", "Nama Pelanggan", "Informasi Trouble", "Jenis Trouble", "Perangkat", "Service",
            "Tanggal Komplain", "Selesai Pengerjaan", "Durasi Pengerjaan", "Keterangan Action",
            "Cluster", "Kategori Cluster", "Created By",
        ]
        data_sheet.append(headers)
        self._apply_header_style(data_sheet)

        for item in rows:
            data_sheet.append([
                item.no_spk,
                item.nama_pelanggan,
                item.informasi_trouble,
                item.jenis_trouble.value,
                item.perangkat.value,
                item.service.value,
                item.tanggal_komplain.strftime("%Y-%m-%d") if item.tanggal_komplain else "",
                item.selesai_pengerjaan.strftime("%Y-%m-%d") if item.selesai_pengerjaan else "",
                item.durasi_pengerjaan if item.durasi_pengerjaan is not None else "",
                item.keterangan_action or "",
                item.cluster_id or "",
                item.kategori_cluster.value if item.kategori_cluster else "",
                item.user.username if hasattr(item, 'user') and item.user else "",
            ])
        self._apply_alternate_rows(data_sheet)
        self._auto_width(data_sheet)

        summary_sheet = wb.create_sheet(title="Ringkasan")
        summary_sheet.append(["Ringkasan Laporan"])
        summary_sheet.append([f"Periode", periode_label])
        summary_sheet.append([])
        summary_sheet.append(["Statistik", "Jumlah"])
        self._apply_header_style(summary_sheet, row=4)
        for key, value in summary.get("kategori", {}).items():
            summary_sheet.append([f"Kategori {key}", value])
        summary_sheet.append([])
        summary_sheet.append(["Jenis Trouble", "Jumlah"])
        self._apply_header_style(summary_sheet, row=summary_sheet.max_row)
        for key, value in summary.get("jenis", {}).items():
            summary_sheet.append([key, value])
        summary_sheet.append([])
        summary_sheet.append(["Perangkat", "Jumlah"])
        self._apply_header_style(summary_sheet, row=summary_sheet.max_row)
        for key, value in summary.get("perangkat", {}).items():
            summary_sheet.append([key, value])
        self._apply_alternate_rows(summary_sheet)
        self._auto_width(summary_sheet)

        info_sheet = wb.create_sheet(title="Clustering Info")
        info_sheet.append(["Informasi Clustering"])
        info_sheet.append(["Tanggal Export", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        info_sheet.append(["Periode", periode_label])
        info_sheet.append(["Model", "K-Means"])
        info_sheet.append(["Random State", 42])
        info_sheet.append(["n_init", 20])
        if clustering_info:
            info_sheet.append([])
            info_sheet.append(["Clustering Terakhir"])
            info_sheet.append(["Tanggal", clustering_info.get("tanggal", "")])
            info_sheet.append(["K", clustering_info.get("k", "")])
            info_sheet.append(["DBI", clustering_info.get("dbi", "")])
            info_sheet.append(["Silhouette", clustering_info.get("silhouette", "")])

        self._apply_alternate_rows(info_sheet)
        self._auto_width(info_sheet)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
